"""Generate a self-contained HTML report from verified stage bundles only."""

from __future__ import annotations

import argparse
from io import BytesIO
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from scripts.utils.contracts import ContractValidationError, load_confirmation_receipt, load_yaml
    from scripts.utils.output_contract import STAGE_SPECS, load_stage_bundle, write_stage_bundle
    from scripts.utils.report_writer import render_template
except ModuleNotFoundError:  # Supports direct execution from the scripts directory.
    from utils.contracts import ContractValidationError, load_confirmation_receipt, load_yaml
    from utils.output_contract import STAGE_SPECS, load_stage_bundle, write_stage_bundle
    from utils.report_writer import render_template


REPORT_SECTIONS = {
    "00": "数据与产品口径",
    "01": "规则挖掘",
    "02": "规则组合",
    "03": "策略效果与 Swap",
}

HTML_CHAPTERS = {
    "overview": "分析概要",
    "data_contract": "数据与口径",
    "rule_mining": "单规则挖掘",
    "rule_combination": "规则组合策略",
    "strategy_swap": "策略效果与 Swap",
    "risk_disclosure": "风险披露",
}

XLSX_SHEETS = [
    "01_报告摘要",
    "02_数据口径",
    "03_特征筛选",
    "04_分箱明细",
    "05_单规则候选",
    "06_Lift排序",
    "07_级联漏斗",
    "08_规则重叠",
    "09_策略效果",
    "10_Swap矩阵",
    "11_风险披露",
]


SHEET_STAGE_MAP = {
    "01_报告摘要": {"00", "01", "02", "03"},
    "02_数据口径": {"00"},
    "03_特征筛选": {"01"},
    "04_分箱明细": {"01"},
    "05_单规则候选": {"01", "02"},
    "06_Lift排序": {"02"},
    "07_级联漏斗": {"02"},
    "08_规则重叠": {"02"},
    "09_策略效果": {"03"},
    "10_Swap矩阵": {"03"},
    "11_风险披露": {"00", "01", "02", "03"},
}


def _source_rows(run_dir: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    rows: list[dict[str, Any]] = []
    validations: list[dict[str, str]] = []
    for stage_id, section_title in REPORT_SECTIONS.items():
        bundle = load_stage_bundle(run_dir, stage_id)
        limitations = "; ".join(bundle["manifest"].get("limitations", [])) or "无"
        for artifact in bundle["inventory"]["artifacts"]:
            rows.append(
                {
                    "section_id": section_title,
                    "stage_id": stage_id,
                    "artifact_name": artifact["artifact_name"],
                    "relative_path": artifact["relative_path"],
                    "sha256": artifact["sha256"],
                    "artifact_status": artifact["status"],
                    "limitation": artifact.get("reason") or limitations,
                }
            )
        validations.append(
            {
                "check_name": f"阶段 {stage_id} 的 manifest、inventory 与产物哈希",
                "status": "passed",
                "details": "已按固定输出契约校验；报告未重新计算指标。",
            }
        )
    return rows, validations


def _write_table(ws, start_row: int, headers: list[str], rows: list[Mapping[str, Any]]) -> None:
    fill = PatternFill("solid", fgColor="F4F7FA")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
    for row_offset, row in enumerate(rows, start=start_row + 1):
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_offset, column=col, value=row.get(header))
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 48)


def _build_xlsx_report(title: str, scope: str, source_rows: list[dict[str, Any]], validations: list[dict[str, str]]) -> bytes:
    """Build the fixed 11-sheet workbook from registered artifact metadata only."""
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    headers = ["artifact_name", "artifact_status", "relative_path", "sha256", "limitation"]
    validation_headers = ["check_name", "status", "details"]
    for sheet_name in XLSX_SHEETS:
        ws = workbook.create_sheet(sheet_name)
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"报告范围：{scope}"
        ws["A3"] = "说明：本工作簿只消费阶段 00-03 已登记且哈希一致的聚合产物；不读取原始数据、不重算指标。"
        ws["A4"] = "观测类型：直接观测、估计、不可观测必须分开展示；缺失指标显示为未提供/不可用。"
        stage_ids = SHEET_STAGE_MAP[sheet_name]
        rows = [
            {key: row.get(key) for key in headers}
            for row in source_rows
            if row.get("stage_id") in stage_ids
        ]
        _write_table(ws, 6, headers, rows)
        if sheet_name == "11_风险披露":
            start = 8 + len(rows)
            _write_table(ws, start, validation_headers, validations)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_final_report(run_dir: str | Path, config: Mapping[str, Any], receipt: Mapping[str, Any]) -> Path:
    """Build stage 04 from registered artifacts, without reading raw source data."""
    source_rows, validations = _source_rows(run_dir)
    title = str(config.get("final_report", {}).get("title") or "信用策略分析报告")
    scope = str(config.get("final_report", {}).get("report_scope") or "以已校验阶段产物为准")
    xlsx_content = _build_xlsx_report(title, scope, source_rows, validations)
    html = render_template(
        Path(__file__).resolve().parents[1] / "templates",
        "final_report.html.j2",
        {
            "title": title,
            "scope": scope,
            "source_rows": source_rows,
            "validations": validations,
            "sections": REPORT_SECTIONS,
            "chapters": HTML_CHAPTERS,
        },
    )
    write_stage_bundle(
        run_dir,
        "04",
        config,
        receipt,
        artifact_rows={
            "report_source_index.csv": source_rows,
            "report_validation.csv": validations,
        },
        binary_artifacts={"final_report.xlsx": xlsx_content},
        limitations=["报告仅汇总已登记且哈希校验通过的聚合产物；不重新计算指标。"],
        html_content=html,
    )
    return Path(run_dir).resolve() / STAGE_SPECS["04"]["directory"] / "final_report.html"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从已校验阶段目录包生成最终 HTML 报告")
    parser.add_argument("--config", required=True, help="UTF-8 YAML 最终报告配置")
    parser.add_argument("--confirmation", required=True, help="UTF-8 JSON 用户确认凭证")
    parser.add_argument("--run-dir", required=True, help="已初始化的运行根目录")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        config = load_yaml(args.config)
        receipt = load_confirmation_receipt(args.confirmation)
        path = build_final_report(args.run_dir, config, receipt)
    except (ContractValidationError, OSError, ValueError) as error:
        print(f"最终报告生成失败: {error}")
        return 2
    print(f"最终报告已写入: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
